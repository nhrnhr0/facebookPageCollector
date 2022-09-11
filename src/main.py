import json
import random
from time import sleep
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
import urllib.parse
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from urllib.parse import urlsplit, parse_qs

from utils import DATA_DIR, check_is_login, get_locations, get_search_keywords

# BASE_PAGES_FACEBOOK_URL = "https://mbasic.facebook.com/search/pages/"
BASE_PAGES_FACEBOOK_URL = "https://facebook.com/search/pages/"


# query = urlsplit(url).query


chrome_options = Options()

# chrome_options.add_argument("--user-data-dir=chrome-data")
if BASE_PAGES_FACEBOOK_URL == "https://mbasic.facebook.com/search/pages/":
    chrome_options.add_experimental_option(
        "prefs", {"profile.managed_default_content_settings.javascript": 2}
    )

driver = webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)
# chrome_options.add_argument("user-data-dir=chrome-data")


def wait_random_time():
    time = random.randint(1, 5)
    sleep(time)


def facebook_login(dr):
    #    driver.get("https://mbasic.facebook.com/login.php")
    url = BASE_PAGES_FACEBOOK_URL.split("search")[0]
    driver.get(url + "login.php")
    # read email and password from file DATA_DIR + "/login.txt"
    # name=email
    # name=pass
    filename = DATA_DIR + "/login.txt"
    wait_random_time()
    with open(filename, "r", encoding="utf-8") as f:
        login = f.readlines()
        login = [l.strip() for l in login]
        email = login[0].split("=")[1]
        password = login[1].split("=")[1]
    # fill email and password

    driver.find_element(by=By.NAME, value="email").send_keys(email)
    driver.find_element(by=By.NAME, value="pass").send_keys(password)
    driver.find_element(by=By.NAME, value="pass").send_keys(Keys.ENTER)
    wait_random_time()
    # check if link לא עכשיו exists and if so click it
    try:
        driver.find_element(
            by=By.XPATH, value="//*[contains(text(), 'לא עכשיו')]"
        ).click()
        wait_random_time()
    except Exception as e:
        pass


driver.get(BASE_PAGES_FACEBOOK_URL)
if check_is_login(driver) == False:
    facebook_login(driver)


def load_page_info(driver, keyword, location_name):
    wait_random_time()
    # 1) look for id="BrowseResultsContainer"
    # 2) find all the tables with role="presentation"
    # 3) get tdoby tr secound td:
    # 4) get a href
    # 5) get all inner divs with and save as dict of key and value (key=class)
    res = []
    # if the driver is location is https://facebook.com...
    if driver.current_url.startswith("https://www.facebook.com"):
        # find data-module-result-type="page"
        try:
            # find all role="article"
            articales = driver.find_elements(
                by=By.CSS_SELECTOR, value="div[role=article]"
            )
            # iterate all the divs inside and click each one, to save the data inside
            for div in articales:
                # click on the div
                # find the a tag inside the div
                a = div.find_element(by=By.TAG_NAME, value="a")
                title = a.accessible_name
                article_text = div.text.split("\n")
                href = a.get_attribute("href")
                res.append(
                    {
                        "title": title,
                        "info": {
                            "href": href,
                            "divs": article_text,
                        },
                    }
                )
        except Exception as e:
            pass
    # if the driver location is https://mbasic.facebook.com...
    elif driver.current_url.startswith("https://mbasic.facebook.com"):
        try:
            result_container = driver.find_element(
                by=By.ID, value="BrowseResultsContainer"
            )
        except Exception as e:
            return res
        tables = result_container.find_elements(by=By.TAG_NAME, value="table")
        for table in tables:
            try:
                tr = table.find_element(by=By.TAG_NAME, value="tr")
                tds = tr.find_elements(by=By.TAG_NAME, value="td")
                td = tds[1]
                a = td.find_element(by=By.TAG_NAME, value="a")
                href = a.get_attribute("href")
                divs = a.find_elements(by=By.TAG_NAME, value="div")
                divs_arr = []

                for div in divs:
                    divs_arr.append(div.text)
                title = divs_arr[0]
                res.append(
                    {
                        "title": title,
                        "info": {
                            "divs": divs_arr,
                            "href": href,
                        },
                    }
                )
            except Exception as e:
                print(e)
                continue
    return res


def request_more_results(driver):
    if BASE_PAGES_FACEBOOK_URL.startswith("https://mbasic.facebook.com"):
        # check for id="see_more_pager"
        # click on it
        try:
            see_more = driver.find_element(by=By.ID, value="see_more_pager")
            # click on the a tag
            see_more.find_element(by=By.TAG_NAME, value="a").click()
            wait_random_time()
            return True
        except Exception as e:

            wait_random_time()
            return False
    elif BASE_PAGES_FACEBOOK_URL.startswith("https://facebook.com"):
        # scroll to the bottom of the page
        # check if there is text "End of results" or סוף התוצאות
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        wait_random_time()
        try:
            end_of_results = driver.find_element(
                by=By.XPATH, value="//*[contains(text(), 'End of results')]"
            )
            return False
        except Exception as e:
            try:
                end_of_results = driver.find_element(
                    by=By.XPATH, value="//*[contains(text(), 'סוף התוצאות')]"
                )
                return False
            except Exception as e:

                # לא מצאנו תוצאות
                try:
                    end_of_results = driver.find_element(
                        by=By.XPATH,
                        value="//*[contains(text(), 'לא מצאנו תוצאות')]",
                    )
                    return False
                except Exception as e:
                    return True


import openpyxl
import os


def is_href_in_ws(ws, href):
    for row in ws.iter_rows(min_row=2):
        tmp = row[3].value
        if tmp == href:
            return True
    return False


def save_data_to_exel(search_data, keyword, location_name):

    # if not create it
    # if exists add data to it (make sure to check the href is not already in the file)
    # columns: keyword, location, title, href, divs
    filename = DATA_DIR + "/data.xlsx"
    # check if exel file DATA_DIR + "/data.xlsx" exists
    if os.path.exists(filename):
        # if exists
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["keyword", "location", "title", "href", "divs"])
    for data in search_data:
        # check if data["info"]["href"] is not already in the file
        # if not add it
        if not is_href_in_ws(ws, data["info"]["href"]):

            ws.append(
                [
                    keyword,
                    location_name,
                    data["title"],
                    data["info"]["href"],
                    "\n".join(data["info"]["divs"]),
                ]
            )

        # ws.append(
        #     [
        #         keyword,
        #         location_name,
        #         data["title"],
        #         data["info"]["href"],
        #         data["info"]["divs"],
        #     ]
        # )
    wb.save(filename)


# https://mbasic.facebook.com/search/pages?q=test
keywords = list(set(get_search_keywords()))
locations = get_locations()
keyword_progress = 0
for keyword in keywords:
    keyword = keyword.strip()
    # escape data
    keyword_url = urllib.parse.quote(keyword)
    location_progress = 0
    for location in locations:
        url = BASE_PAGES_FACEBOOK_URL + "?q=" + keyword_url + "&" + location["location"]
        driver.get(url)
        search_data = []
        location_progress += 1
        while True:
            pages = load_page_info(driver, keyword, location["name"])
            search_data += pages
            if not request_more_results(driver):
                break
        # write data to file data.exel
        save_data_to_exel(search_data, keyword, location["name"])

        print(
            "keyword progress: {}/{} {} location progress: {}/{} {}".format(
                keyword_progress,
                len(keywords),
                keyword,
                location_progress,
                len(locations),
                location["name"],
            )
        )
    keyword_progress += 1
    wait_random_time()
sleep(3550)
driver.close()
